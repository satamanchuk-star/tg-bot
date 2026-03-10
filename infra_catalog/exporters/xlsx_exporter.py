"""Экспорт в XLSX."""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from ..config import CENTER_LAT, CENTER_LON, SEARCH_RADIUS_KM
from ..constants import CATEGORY_SUBCATEGORY
from ..models import InfraObject

logger = logging.getLogger(__name__)

_FIELDS = [
    "name", "category", "subcategory", "address", "phone", "website",
    "work_time", "description", "lat", "lon", "distance_km", "source", "is_active",
]


def export_xlsx(objects: list[InfraObject], path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.error("openpyxl не установлен, XLSX-экспорт пропущен")
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Лист objects
    ws = wb.active
    ws.title = "objects"
    ws.append(_FIELDS)
    for obj in objects:
        row = obj.model_dump()
        row["distance_km"] = round(row["distance_km"], 3)
        ws.append([row.get(f, "") for f in _FIELDS])

    # Лист category_dictionary
    ws_dict = wb.create_sheet("category_dictionary")
    ws_dict.append(["category", "subcategory"])
    for cat, subs in CATEGORY_SUBCATEGORY.items():
        for sub in subs:
            ws_dict.append([cat, sub])

    # Лист meta
    active_count = sum(1 for o in objects if o.is_active)
    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["key", "value"])
    ws_meta.append(["center_lat", CENTER_LAT])
    ws_meta.append(["center_lon", CENTER_LON])
    ws_meta.append(["radius_km", SEARCH_RADIUS_KM])
    ws_meta.append(["generated_at", datetime.now(timezone.utc).isoformat()])
    ws_meta.append(["total_objects", len(objects)])
    ws_meta.append(["active_objects", active_count])
    ws_meta.append(["inactive_objects", len(objects) - active_count])

    wb.save(path)
    logger.info("XLSX: %d объектов -> %s", len(objects), path)
