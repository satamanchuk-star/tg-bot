"""Почему: банк вопросов викторины ведётся владельцем в XLSX (колонки
«Вопрос | Ответ»). В ответе сначала идёт короткий ответ, затем пояснение —
конвертер разделяет их: короткое — для матча, пояснение — показать после
развязки вопроса.

Форматы ответа, которые понимает разбор:
- «Кинотеатры. 5 центов стоил билет…»        → ответ «Кинотеатры», пояснение дальше
- «Собаки-поводыри (зачёт: рыбы-поводыри) …» → альтернатива через « / »
- «Да (Yes). …»                              → короткая скобка = альтернатива
- «Табун (обыгрывается сходство…)»           → длинная скобка = пояснение

Использование: python -m scripts.import_quiz_xlsx <файл.xlsx>
Пишет data/quiz_questions.json (только валидные пары) и печатает статистику.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

OUT_FILE = Path(__file__).resolve().parent.parent / "data" / "quiz_questions.json"

# Скобка внутри первого сегмента ответа.
_PAREN_RE = re.compile(r"\(([^)]*)\)")
# Первое предложение (до «. » / «! » / «? » с последующим текстом).
_SENT_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")


def split_answer(raw: str) -> tuple[str, str]:
    """Разделяет сырой ответ на (короткий ответ для матча, пояснение).

    Короткий ответ может содержать альтернативы через « / » (из «зачёт: …»
    и коротких скобок). Пояснение — всё остальное, показывается при развязке.
    """
    text = " ".join(str(raw).split()).strip()
    if not text:
        return "", ""

    parts = _SENT_SPLIT_RE.split(text, maxsplit=1)
    first = parts[0].strip()
    comment = parts[1].strip() if len(parts) > 1 else ""

    # Скобки в первом сегменте: «зачёт: X» и короткие (≤2 слов) — альтернативы,
    # длинные — уходят в пояснение.
    alternatives: list[str] = []
    extra_comments: list[str] = []

    def _handle_paren(m: re.Match) -> str:
        inner = m.group(1).strip()
        low = inner.lower()
        if low.startswith(("зачёт", "зачет")):
            alt = inner.split(":", 1)[1].strip() if ":" in inner else ""
            if alt:
                alternatives.append(alt)
        elif inner and len(inner.split()) <= 2 and "–" not in inner and "—" not in inner:
            alternatives.append(inner)
        elif inner:
            extra_comments.append(inner)
        return " "

    short = _PAREN_RE.sub(_handle_paren, first)
    short = " ".join(short.split()).strip().rstrip(".!?").strip()

    if alternatives:
        short = " / ".join([short] + [a.rstrip(".!?") for a in alternatives if a])
    if extra_comments:
        comment = " ".join(extra_comments + ([comment] if comment else []))
    return short, comment


def convert(xlsx_path: str) -> tuple[list[dict], list[str]]:
    """Читает XLSX и возвращает (валидные записи, причины отбраковки)."""
    import openpyxl

    from scripts.validate_quiz import validate_one

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    items: list[dict] = []
    rejected: list[str] = []
    seen_questions: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        question = " ".join(str(row[0]).split()).strip()
        answer_short, comment = split_answer(str(row[1]))
        if not answer_short:
            rejected.append(f"«{question[:60]}»: пустой ответ после разбора")
            continue
        norm_q = " ".join(re.sub(r"[^\w\s]", " ", question.lower().replace("ё", "е")).split())
        if norm_q in seen_questions:
            rejected.append(f"«{question[:60]}»: дубль вопроса")
            continue
        item = {"question": question, "answer": answer_short, "category": "квиз"}
        if comment:
            item["comment"] = comment[:500]
        issues = validate_one(item)
        if issues:
            rejected.append(f"«{question[:60]}» → «{answer_short[:40]}»: {issues[0]}")
            continue
        seen_questions.add(norm_q)
        items.append(item)

    return items, rejected


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Использование: python -m scripts.import_quiz_xlsx <файл.xlsx>", file=sys.stderr)
        raise SystemExit(2)
    items, rejected = convert(sys.argv[1])
    OUT_FILE.write_text(
        json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"валидных: {len(items)}, отбраковано: {len(rejected)} → {OUT_FILE}", file=sys.stderr)
    for line in rejected[:30]:
        print("  отсеяно:", line, file=sys.stderr)